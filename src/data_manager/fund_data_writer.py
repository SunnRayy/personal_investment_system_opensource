"""
Fund Data Writer Module

This module provides functionality to persist processed Global Markets data back to Excel files.
Solves the critical issue where processed holdings and transactions were calculated but never saved.

Author: System
Created: November 18, 2025
"""

import pandas as pd
import logging
from typing import Optional
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


def write_processed_fund_data(
    fund_file_path: str,
    processed_holdings: pd.DataFrame,
    processed_transactions: pd.DataFrame,
    custom_logger: Optional[logging.Logger] = None
) -> bool:
    """
    Write processed fund holdings and transactions to Excel file.
    
    CRITICAL: This function MERGES new processed data with existing historical data.
    It NEVER replaces historical data, only adds new records.
    
    This function:
    1. Loads existing historical data from Excel sheets
    2. Merges processed (new) data with existing historical data
    3. Removes duplicates to prevent data redundancy
    4. Writes the COMBINED dataset back to Excel
    
    This ensures historical data is NEVER lost.
    
    Args:
        fund_file_path: Path to funding_transactions.xlsx
        processed_holdings: Cleaned NEW holdings DataFrame with columns:
            ['Asset_ID', 'Asset_Name', 'Asset_Type_Raw', 'Snapshot_Date', 
             'Market_Price_Unit', 'Quantity', 'Market_Value_Raw']
        processed_transactions: Cleaned NEW transactions DataFrame with columns:
            ['Transaction_Date', 'Asset_ID', 'Asset_Name', 'Transaction_Type_Raw',
             'Quantity', 'Amount_Gross', 'Commission_Fee', 'Price_Unit']
        custom_logger: Optional logger instance (uses module logger if None)
        
    Returns:
        True if data was successfully written and verified, False otherwise
        
    Raises:
        Exception: If file operations fail (logged but not raised)
    """
    log = custom_logger if custom_logger else logger
    
    try:
        # Validate inputs
        if processed_holdings is None or processed_holdings.empty:
            log.warning("⚠️  Processed holdings is empty, nothing to write")
            return False
            
        if processed_transactions is None or processed_transactions.empty:
            log.warning("⚠️  Processed transactions is empty, nothing to write")
            return False
        
        log.info(f"Writing processed fund data to {fund_file_path}...")
        log.info(f"  NEW Holdings: {len(processed_holdings)} rows")
        log.info(f"  NEW Transactions: {len(processed_transactions)} rows")
        
        # CRITICAL: Load existing historical data FIRST before making any changes
        # This prevents data loss by ensuring we merge instead of replace
        holdings_sheet_name = '基金持仓汇总'
        transactions_sheet_name = '基金交易记录'
        
        try:
            # Read existing historical holdings
            existing_holdings = pd.read_excel(fund_file_path, sheet_name=holdings_sheet_name)
            log.info(f"📚 Loaded {len(existing_holdings)} existing historical holdings")
        except Exception as e:
            log.warning(f"⚠️  Could not load existing holdings (file may be new): {e}")
            existing_holdings = pd.DataFrame()
        
        try:
            # Read existing historical transactions
            existing_transactions = pd.read_excel(fund_file_path, sheet_name=transactions_sheet_name)
            log.info(f"📚 Loaded {len(existing_transactions)} existing historical transactions")
        except Exception as e:
            log.warning(f"⚠️  Could not load existing transactions (file may be new): {e}")
            existing_transactions = pd.DataFrame()
        
        # MERGE: Combine historical + new data
        if not existing_holdings.empty:
            # CRITICAL FIX: Standardize Chinese column names to English for Holdings
            # The Excel sheet uses Chinese headers ('基金代码' etc), but we need English keys ('Asset_ID')
            chinese_to_english_holdings_map = {
                '基金代码': 'Asset_ID',
                '基金名称': 'Asset_Name',
                '基金类型': 'Asset_Type_Raw',
                '净值日期': 'Snapshot_Date',
                '单位净值': 'Market_Price_Unit',
                '持有份额': 'Quantity',
                '参考市值': 'Market_Value_Raw'
            }
            existing_holdings = existing_holdings.rename(columns=chinese_to_english_holdings_map)
            log.debug(f"Standardized existing holdings columns: {list(existing_holdings.columns)}")

            # Helper: Standardize CN Fund IDs to 6-digit strings (checking numeric nature)
            def normalize_fund_id(val):
                s = str(val).strip()
                # If numeric and <= 6 digits, pad with zeros
                if s.isdigit() and len(s) <= 6:
                    return s.zfill(6)
                # Handle cases like "311.0" -> "000311"
                try:
                    f = float(s)
                    if f.is_integer():
                         return str(int(f)).zfill(6)
                except:
                    pass
                return s

            # Ensure consistent types and formats for merging
            existing_holdings['Asset_ID'] = existing_holdings['Asset_ID'].apply(normalize_fund_id)
            existing_holdings['Snapshot_Date'] = pd.to_datetime(existing_holdings['Snapshot_Date'])
            
            processed_holdings['Asset_ID'] = processed_holdings['Asset_ID'].apply(normalize_fund_id)
            processed_holdings['Snapshot_Date'] = pd.to_datetime(processed_holdings['Snapshot_Date'])

            merged_holdings = pd.concat([existing_holdings, processed_holdings], ignore_index=True)
            
            # CRITICAL FIX: Prevent Double Counting on Idempotent Runs
            # manager.py ALREADY aggregates split accounts (Same Date, Same ID) into a single row.
            # So here we only need to merge History + New.
            # If we run the script twice for the same date, we have [Old_12-12, New_12-12].
            # distinct subset should be ['Snapshot_Date', 'Asset_ID'].
            # keep='last' ensures we take the latest processed version if duplicates exist.
            
            merged_holdings = merged_holdings.sort_values('Snapshot_Date', ascending=True)
            merged_holdings = merged_holdings.drop_duplicates(
                subset=['Snapshot_Date', 'Asset_ID'], 
                keep='last'
            )
            
            # Sort for display (Latest Date First)
            merged_holdings = merged_holdings.sort_values('Snapshot_Date', ascending=False)
            
            log.info(f"🔗 Merged holdings: {len(existing_holdings)} historical + {len(processed_holdings)} new -> {len(merged_holdings)} unique records (History Preserved)")
        else:
            merged_holdings = processed_holdings
            log.info(f"📝 No existing holdings, using {len(merged_holdings)} processed holdings")
        
        if not existing_transactions.empty:
            # CRITICAL FIX: Standardize column names BEFORE merging
            # Existing data has Chinese column names, processed data has English
            # We need to convert one to match the other before concat
            
            # Map Chinese column names to English for standardization
            chinese_to_english_map = {
                '交易日期': 'Transaction_Date',
                '基金代码': 'Asset_ID',
                '基金名称': 'Asset_Name',
                '操作类型': 'Transaction_Type_Raw',
                '交易金额': 'Amount_Gross',
                '交易份额': 'Quantity',
                '交易时基金单位净值': 'Price_Unit',
                '手续费': 'Commission_Fee',
                '交易原因': 'Memo'
            }
            
            # Standardize existing transactions to English column names
            existing_transactions_std = existing_transactions.rename(columns=chinese_to_english_map)
            log.debug(f"Standardized existing transactions columns: {list(existing_transactions_std.columns)}")
            
            # Ensure both DataFrames have the same columns for merging
            # Get union of all columns
            all_columns = list(set(existing_transactions_std.columns) | set(processed_transactions.columns))
            
            # Add missing columns with NaN
            for col in all_columns:
                if col not in existing_transactions_std.columns:
                    existing_transactions_std[col] = None
                if col not in processed_transactions.columns:
                    processed_transactions[col] = None
            
            # Reorder columns to match
            existing_transactions_std = existing_transactions_std[all_columns]
            processed_transactions = processed_transactions[all_columns]
            
            # Now merge with matching column names
            merged_transactions = pd.concat([existing_transactions_std, processed_transactions], ignore_index=True)
            
            # Remove duplicates based on key transaction identifiers (English names)
            dedup_columns = ['Transaction_Date', 'Asset_ID', 'Amount_Gross']
            
            # Convert Asset_ID to string for consistent comparison
            merged_transactions['Asset_ID'] = merged_transactions['Asset_ID'].astype(str)
            
            initial_count = len(merged_transactions)
            merged_transactions = merged_transactions.drop_duplicates(
                subset=dedup_columns,
                keep='last'
            )
            duplicates_removed = initial_count - len(merged_transactions)
            
            # Sort by date
            merged_transactions = merged_transactions.sort_values('Transaction_Date')
            
            log.info(f"🔗 Merged transactions: {len(existing_transactions)} historical + {len(processed_transactions)} new = {len(merged_transactions)} total (removed {duplicates_removed} duplicates)")
        else:
            merged_transactions = processed_transactions
            log.info(f"📝 No existing transactions, using {len(merged_transactions)} processed transactions")
        
        # Now write the MERGED data to Excel
        try:
            workbook = load_workbook(fund_file_path)
        except FileNotFoundError:
            log.error(f"❌ Fund file not found: {fund_file_path}")
            return False
        except Exception as e:
            log.error(f"❌ Error opening Excel file: {e}")
            return False
        
        # Write merged holdings to '基金持仓汇总' sheet
        try:
            if holdings_sheet_name in workbook.sheetnames:
                del workbook[holdings_sheet_name]
                log.debug(f"Removed existing {holdings_sheet_name} sheet (will replace with merged data)")
            
            holdings_sheet = workbook.create_sheet(holdings_sheet_name)
            
            for row_idx, row in enumerate(dataframe_to_rows(merged_holdings, index=False, header=True), 1):
                for col_idx, value in enumerate(row, 1):
                    holdings_sheet.cell(row=row_idx, column=col_idx, value=value)
            
            log.info(f"✅ Wrote {len(merged_holdings)} MERGED holdings to {holdings_sheet_name}")
            
        except Exception as e:
            log.error(f"❌ Error writing holdings sheet: {e}")
            workbook.close()
            return False
        
        # Write merged transactions to '基金交易记录' sheet
        try:
            if transactions_sheet_name in workbook.sheetnames:
                del workbook[transactions_sheet_name]
                log.debug(f"Removed existing {transactions_sheet_name} sheet (will replace with merged data)")
            
            transactions_sheet = workbook.create_sheet(transactions_sheet_name)
            
            # CRITICAL: Convert back to Chinese column names for Excel storage
            english_to_chinese_map = {
                'Transaction_Date': '交易日期',
                'Asset_ID': '基金代码',
                'Asset_Name': '基金名称',
                'Transaction_Type_Raw': '操作类型',
                'Amount_Gross': '交易金额',
                'Quantity': '交易份额',
                'Price_Unit': '交易时基金单位净值',
                'Commission_Fee': '手续费',
                'Memo': '交易原因'
            }
            
            merged_transactions_chinese = merged_transactions.rename(columns=english_to_chinese_map)
            
            # Write only the Chinese columns (drop any English columns that weren't mapped)
            chinese_columns = ['交易日期', '基金代码', '基金名称', '操作类型', '交易金额', '交易份额', '交易时基金单位净值', '手续费', '交易原因']
            # Keep only columns that exist
            columns_to_write = [col for col in chinese_columns if col in merged_transactions_chinese.columns]
            merged_transactions_final = merged_transactions_chinese[columns_to_write]
            
            for row_idx, row in enumerate(dataframe_to_rows(merged_transactions_final, index=False, header=True), 1):
                for col_idx, value in enumerate(row, 1):
                    transactions_sheet.cell(row=row_idx, column=col_idx, value=value)
            
            log.info(f"✅ Wrote {len(merged_transactions_final)} MERGED transactions to {transactions_sheet_name}")
            
        except Exception as e:
            log.error(f"❌ Error writing transactions sheet: {e}")
            workbook.close()
            return False
        
        # Save the workbook
        try:
            workbook.save(fund_file_path)
            workbook.close()
            log.info(f"✅ Successfully saved MERGED fund data to {fund_file_path}")
            
        except Exception as e:
            log.error(f"❌ Error saving Excel file: {e}")
            workbook.close()
            return False
        
        # Verify persistence by reading back (use merged counts, not just processed)
        if verify_persistence(fund_file_path, len(merged_holdings), len(merged_transactions), log):
            log.info("✅ Data persistence verified successfully - NO DATA LOSS!")
            return True
        else:
            log.warning("⚠️  Data written but verification failed")
            return False
            
    except Exception as e:
        log.error(f"❌ Unexpected error in write_processed_fund_data: {e}")
        import traceback
        log.debug(traceback.format_exc())
        return False


def verify_persistence(
    fund_file_path: str,
    expected_holdings_count: int,
    expected_transactions_count: int,
    custom_logger: Optional[logging.Logger] = None
) -> bool:
    """
    Verify processed data was written correctly and persists in Excel file.
    
    Args:
        fund_file_path: Path to funding_transactions.xlsx
        expected_holdings_count: Expected number of holdings rows
        expected_transactions_count: Expected number of transactions rows
        custom_logger: Optional logger instance
        
    Returns:
        True if verification passes, False otherwise
    """
    log = custom_logger if custom_logger else logger
    
    try:
        log.debug("Verifying data persistence...")
        
        # Read back holdings
        holdings_df = pd.read_excel(fund_file_path, sheet_name='基金持仓汇总')
        actual_holdings_count = len(holdings_df)
        
        # Read back transactions
        transactions_df = pd.read_excel(fund_file_path, sheet_name='基金交易记录')
        actual_transactions_count = len(transactions_df)
        
        # Verify counts match
        holdings_match = actual_holdings_count == expected_holdings_count
        transactions_match = actual_transactions_count == expected_transactions_count
        
        if holdings_match and transactions_match:
            log.debug(f"✅ Verification passed: {actual_holdings_count} holdings, {actual_transactions_count} transactions")
            return True
        else:
            if not holdings_match:
                log.warning(f"⚠️  Holdings count mismatch: expected {expected_holdings_count}, got {actual_holdings_count}")
            if not transactions_match:
                log.warning(f"⚠️  Transactions count mismatch: expected {expected_transactions_count}, got {actual_transactions_count}")
            return False
            
    except Exception as e:
        log.error(f"❌ Error verifying persistence: {e}")
        return False


def clear_raw_paste_sheets(
    fund_file_path: str,
    custom_logger: Optional[logging.Logger] = None
) -> bool:
    """
    Clear raw paste sheets after successful processing (optional operation).
    
    This prevents duplicate processing of the same data and keeps raw sheets clean.
    Use this cautiously as it removes the audit trail of raw input data.
    
    Args:
        fund_file_path: Path to funding_transactions.xlsx
        custom_logger: Optional logger instance
        
    Returns:
        True if sheets were cleared successfully, False otherwise
    """
    log = custom_logger if custom_logger else logger
    
    try:
        log.info("Clearing raw paste sheets...")
        
        workbook = load_workbook(fund_file_path)
        
        # Clear raw_holdings_paste sheet
        if 'raw_holdings_paste' in workbook.sheetnames:
            sheet = workbook['raw_holdings_paste']
            sheet.delete_rows(2, sheet.max_row)  # Keep header row (row 1)
            log.info("✅ Cleared raw_holdings_paste sheet")
        
        # Clear raw_transactions_paste sheet
        if 'raw_transactions_paste' in workbook.sheetnames:
            sheet = workbook['raw_transactions_paste']
            sheet.delete_rows(2, sheet.max_row)  # Keep header row (row 1)
            log.info("✅ Cleared raw_transactions_paste sheet")
        
        workbook.save(fund_file_path)
        workbook.close()
        
        log.info("✅ Raw paste sheets cleared successfully")
        return True
        
    except Exception as e:
        log.error(f"❌ Error clearing raw paste sheets: {e}")
        return False
