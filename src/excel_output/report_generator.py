"""
Excel Report Generator for Personal Investment Analysis

This module provides the main orchestrator for generating comprehensive
Excel reports from financial analysis and recommendation data.
"""

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Union
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Excel output imports
try:
    from .worksheet_manager import WorksheetManager
    from .formatters import ExcelFormatter
    from .chart_builder import ChartBuilder
except ImportError:
    # Fallback for standalone execution
    WorksheetManager = None
    ExcelFormatter = None
    ChartBuilder = None

# Recommendation engine imports
try:
    from recommendation_engine.comprehensive_engine import ComprehensiveRecommendationResult
    from recommendation_engine.action_prioritizer import ActionPlan
except ImportError:
    # Fallback for different import contexts
    try:
        from src.recommendation_engine.comprehensive_engine import ComprehensiveRecommendationResult
        from src.recommendation_engine.action_prioritizer import ActionPlan
    except ImportError:
        # Define minimal stubs if imports fail
        ComprehensiveRecommendationResult = None
        ActionPlan = None

logger = logging.getLogger(__name__)


class ExcelReportGenerator:
    """
    Main Excel Report Generator
    
    Orchestrates the creation of comprehensive Excel reports from
    financial analysis and recommendation data.
    """
    
    def __init__(self, output_dir: str = "output"):
        """
        Initialize Excel Report Generator
        
        Args:
            output_dir: Directory for saving Excel reports
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        
        self.worksheet_manager = WorksheetManager()
        self.formatter = ExcelFormatter()
        self.chart_builder = ChartBuilder()
        
        logger.info("Excel Report Generator initialized")
    
    def generate_comprehensive_report(
        self,
        recommendation_result,
        financial_data: Optional[Dict] = None,
        portfolio_data: Optional[Dict] = None,
        filename: Optional[str] = None
    ) -> str:
        """
        Generate comprehensive Excel report
        
        Args:
            recommendation_result: Comprehensive recommendation data
            financial_data: Financial analysis data
            portfolio_data: Portfolio analysis data
            filename: Optional custom filename
        
        Returns:
            Path to generated Excel file
        """
        logger.info("Starting comprehensive Excel report generation")
        
        # Validate input
        if not recommendation_result:
            raise ValueError("recommendation_result is required")
        
        # Create workbook
        workbook = Workbook()
        
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"investment_analysis_report_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        try:
            # Create worksheets with error handling
            logger.info("Creating executive summary...")
            self._create_executive_summary(workbook, recommendation_result)
            
            logger.info("Creating recommendations detail...")
            self._create_recommendations_detail(workbook, recommendation_result)
            
            logger.info("Creating action plan...")
            self._create_action_plan(workbook, recommendation_result)
            
            logger.info("Creating risk analysis...")
            self._create_risk_analysis(workbook, recommendation_result)
            
            if financial_data:
                logger.info("Creating financial summary...")
                self._create_financial_summary(workbook, financial_data)
            
            if portfolio_data:
                logger.info("Creating portfolio summary...")
                self._create_portfolio_summary(workbook, portfolio_data)
            
            # Remove default worksheet
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])
            
            # Save workbook
            workbook.save(filepath)
            logger.info(f"Excel report saved to: {filepath}")
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            raise
    
    def _create_executive_summary(
        self, 
        workbook: Workbook,
        result
    ):
        """Create executive summary worksheet"""
        ws = workbook.create_sheet("Executive Summary", 0)
        
        # Title
        ws['A1'] = "Investment Analysis - Executive Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        
        # Summary metrics
        row = 3
        ws[f'A{row}'] = "Report Generated:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        row += 1
        ws[f'A{row}'] = "Total Recommendations:"
        total_recs = (len(result.financial_recommendations) + 
                     len(result.portfolio_recommendations) + 
                     len(result.risk_recommendations))
        ws[f'B{row}'] = total_recs
        
        row += 1
        ws[f'A{row}'] = "High Priority Actions:"
        
        # Get all actions from action plan with validation
        all_actions = []
        high_priority = 0
        
        try:
            if hasattr(result, 'prioritized_action_plan') and result.prioritized_action_plan:
                action_plan = result.prioritized_action_plan
                
                # Safely collect actions from different categories
                if hasattr(action_plan, 'immediate_actions') and action_plan.immediate_actions:
                    all_actions.extend(action_plan.immediate_actions)
                if hasattr(action_plan, 'short_term_actions') and action_plan.short_term_actions:
                    all_actions.extend(action_plan.short_term_actions)
                if hasattr(action_plan, 'medium_term_actions') and action_plan.medium_term_actions:
                    all_actions.extend(action_plan.medium_term_actions)
                if hasattr(action_plan, 'long_term_actions') and action_plan.long_term_actions:
                    all_actions.extend(action_plan.long_term_actions)
                    
                high_priority = len([action for action in all_actions 
                                   if hasattr(action, 'priority') and action.priority == "HIGH"])
            else:
                logger.warning("No prioritized action plan found in recommendation result")
        except Exception as e:
            logger.error(f"Error processing action plan in executive summary: {e}")
            
        ws[f'B{row}'] = high_priority
        
        # Key insights
        row += 3
        ws[f'A{row}'] = "Key Insights"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        
        # Add top recommendations summary
        row += 2
        if result.financial_recommendations:
            ws[f'A{row}'] = "• Financial: " + result.financial_recommendations[0].description[:50] + "..."
            row += 1
        
        if result.portfolio_recommendations:
            ws[f'A{row}'] = "• Portfolio: " + result.portfolio_recommendations[0].description[:50] + "..."
            row += 1
            
        if result.risk_recommendations:
            ws[f'A{row}'] = "• Risk: " + result.risk_recommendations[0].description[:50] + "..."
        
        # Format the worksheet
        self.formatter.format_summary_sheet(ws)
    
    def _create_recommendations_detail(
        self,
        workbook: Workbook,
        result
    ):
        """Create detailed recommendations worksheet"""
        ws = workbook.create_sheet("Recommendations Detail")
        
        # Headers
        headers = ["Category", "Priority", "Description", "Expected Impact", "Timeline"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Add recommendations
        row = 2
        
        # Financial recommendations
        for rec in result.financial_recommendations:
            ws.cell(row=row, column=1, value="Financial")
            ws.cell(row=row, column=2, value=rec.priority)
            ws.cell(row=row, column=3, value=rec.description)
            ws.cell(row=row, column=4, value=rec.expected_impact)
            ws.cell(row=row, column=5, value=rec.timeline)
            row += 1
        
        # Portfolio recommendations
        for rec in result.portfolio_recommendations:
            ws.cell(row=row, column=1, value="Portfolio")
            ws.cell(row=row, column=2, value=rec.priority)
            ws.cell(row=row, column=3, value=rec.description)
            ws.cell(row=row, column=4, value=rec.expected_impact)
            ws.cell(row=row, column=5, value=rec.timeline)
            row += 1
        
        # Risk recommendations
        for rec in result.risk_recommendations:
            ws.cell(row=row, column=1, value="Risk")
            ws.cell(row=row, column=2, value=rec.priority)
            ws.cell(row=row, column=3, value=rec.description)
            ws.cell(row=row, column=4, value=rec.expected_impact)
            ws.cell(row=row, column=5, value=rec.timeline)
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        self.formatter.format_data_sheet(ws)
    
    def _create_action_plan(
        self,
        workbook: Workbook,
        result
    ):
        """Create action plan worksheet"""
        ws = workbook.create_sheet("Action Plan")
        
        # Headers
        headers = ["Priority", "Action", "Category", "Timeline", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Add actions with proper validation
        row = 2
        all_actions = []
        
        try:
            if hasattr(result, 'prioritized_action_plan') and result.prioritized_action_plan:
                action_plan = result.prioritized_action_plan
                
                # Collect all actions from different categories with safe access
                if hasattr(action_plan, 'immediate_actions') and action_plan.immediate_actions:
                    all_actions.extend(action_plan.immediate_actions)
                if hasattr(action_plan, 'short_term_actions') and action_plan.short_term_actions:
                    all_actions.extend(action_plan.short_term_actions)
                if hasattr(action_plan, 'medium_term_actions') and action_plan.medium_term_actions:
                    all_actions.extend(action_plan.medium_term_actions)
                if hasattr(action_plan, 'long_term_actions') and action_plan.long_term_actions:
                    all_actions.extend(action_plan.long_term_actions)
                if hasattr(action_plan, 'monitoring_actions') and action_plan.monitoring_actions:
                    all_actions.extend(action_plan.monitoring_actions)
            else:
                logger.warning("No prioritized action plan found in recommendation result")
                # Add a placeholder row
                ws.cell(row=row, column=1, value="INFO")
                ws.cell(row=row, column=2, value="No action plan available")
                ws.cell(row=row, column=3, value="System")
                ws.cell(row=row, column=4, value="N/A")
                ws.cell(row=row, column=5, value="N/A")
                row += 1
        except Exception as e:
            logger.error(f"Error processing action plan: {e}")
            # Add error row
            ws.cell(row=row, column=1, value="ERROR")
            ws.cell(row=row, column=2, value=f"Error processing action plan: {str(e)}")
            ws.cell(row=row, column=3, value="System")
            ws.cell(row=row, column=4, value="N/A")
            ws.cell(row=row, column=5, value="N/A")
            row += 1
        
        # Add actions to worksheet
        for action in all_actions:
            try:
                ws.cell(row=row, column=1, value=getattr(action, 'priority', 'MEDIUM'))
                ws.cell(row=row, column=2, value=getattr(action, 'description', str(action)))
                ws.cell(row=row, column=3, value=getattr(action, 'category', 'General'))
                ws.cell(row=row, column=4, value=getattr(action, 'timeline', 'TBD'))
                ws.cell(row=row, column=5, value="Pending")
                row += 1
            except Exception as e:
                logger.error(f"Error processing individual action: {e}")
                continue
        
        # Format priority cells with colors
        for row_num in range(2, row):
            priority_cell = ws.cell(row=row_num, column=1)
            if priority_cell.value == "HIGH":
                priority_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif priority_cell.value == "MEDIUM":
                priority_cell.fill = PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid")
            elif priority_cell.value == "LOW":
                priority_cell.fill = PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        self.formatter.format_data_sheet(ws)
    
    def _create_risk_analysis(
        self,
        workbook: Workbook,
        result
    ):
        """Create risk analysis worksheet"""
        ws = workbook.create_sheet("Risk Analysis")
        
        # Title
        ws['A1'] = "Risk Assessment Summary"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Risk metrics (if available)
        row = 3
        ws[f'A{row}'] = "Risk Recommendations Count:"
        ws[f'B{row}'] = len(result.risk_recommendations)
        
        row += 2
        ws[f'A{row}'] = "Risk Recommendations"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        
        # Headers for risk recommendations
        row += 1
        headers = ["Priority", "Risk Type", "Description", "Mitigation"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Add risk recommendations
        row += 1
        for rec in result.risk_recommendations:
            ws.cell(row=row, column=1, value=rec.priority)
            ws.cell(row=row, column=2, value="Risk Assessment")
            ws.cell(row=row, column=3, value=rec.description)
            ws.cell(row=row, column=4, value=rec.expected_impact)
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        self.formatter.format_data_sheet(ws)
    
    def _create_financial_summary(self, workbook: Workbook, financial_data: Dict):
        """Create financial summary worksheet"""
        ws = workbook.create_sheet("Financial Summary")
        
        # This would be expanded based on actual financial data structure
        ws['A1'] = "Financial Analysis Summary"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Placeholder for financial data
        ws['A3'] = "Financial data integration pending..."
        
        self.formatter.format_summary_sheet(ws)
    
    def _create_portfolio_summary(self, workbook: Workbook, portfolio_data: Dict):
        """Create portfolio summary worksheet"""
        ws = workbook.create_sheet("Portfolio Summary")
        
        # This would be expanded based on actual portfolio data structure
        ws['A1'] = "Portfolio Analysis Summary"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Placeholder for portfolio data
        ws['A3'] = "Portfolio data integration pending..."
        
        self.formatter.format_summary_sheet(ws)
