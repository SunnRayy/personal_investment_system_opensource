"""
Worksheet Manager for Personal Investment Analysis

This module provides worksheet management capabilities
for organizing Excel report structure and content.
"""

import logging
from typing import Dict, List, Optional
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class WorksheetManager:
    """
    Worksheet Manager for Excel Reports
    
    Manages the creation, organization, and structure of Excel worksheets
    for comprehensive investment analysis reports.
    """
    
    def __init__(self):
        """Initialize Worksheet Manager"""
        self.worksheet_config = {
            "Executive Summary": {
                "position": 0,
                "tab_color": "366092"
            },
            "Recommendations Detail": {
                "position": 1,
                "tab_color": "4472C4"
            },
            "Action Plan": {
                "position": 2,
                "tab_color": "70AD47"
            },
            "Risk Analysis": {
                "position": 3,
                "tab_color": "E74C3C"
            },
            "Financial Summary": {
                "position": 4,
                "tab_color": "FFC000"
            },
            "Portfolio Summary": {
                "position": 5,
                "tab_color": "7030A0"
            }
        }
        
        logger.info("Worksheet Manager initialized")
    
    def create_worksheet_structure(self, workbook: Workbook) -> Dict[str, Worksheet]:
        """
        Create the standard worksheet structure
        
        Args:
            workbook: Excel workbook to organize
            
        Returns:
            Dictionary mapping worksheet names to worksheet objects
        """
        worksheets = {}
        
        try:
            # Remove default sheet if it exists
            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])
            
            # Create worksheets in order
            for name, config in self.worksheet_config.items():
                ws = workbook.create_sheet(name, config["position"])
                ws.sheet_properties.tabColor = config["tab_color"]
                worksheets[name] = ws
                logger.debug(f"Created worksheet: {name}")
            
            logger.info("Worksheet structure created successfully")
            return worksheets
            
        except Exception as e:
            logger.error(f"Error creating worksheet structure: {e}")
            raise
    
    def organize_worksheets(self, workbook: Workbook):
        """
        Organize existing worksheets according to standard order
        
        Args:
            workbook: Excel workbook to organize
        """
        try:
            # Get current worksheets
            current_sheets = workbook.sheetnames.copy()
            
            # Reorder according to configuration
            for name, config in self.worksheet_config.items():
                if name in current_sheets:
                    ws = workbook[name]
                    workbook.move_sheet(ws, config["position"])
                    ws.sheet_properties.tabColor = config["tab_color"]
            
            logger.info("Worksheets organized successfully")
            
        except Exception as e:
            logger.error(f"Error organizing worksheets: {e}")
            raise
    
    def add_custom_worksheet(
        self, 
        workbook: Workbook, 
        name: str, 
        position: Optional[int] = None,
        tab_color: str = "808080"
    ) -> Worksheet:
        """
        Add a custom worksheet to the workbook
        
        Args:
            workbook: Excel workbook
            name: Name for the new worksheet
            position: Position index (None for end)
            tab_color: Hex color code for tab
            
        Returns:
            Created worksheet
        """
        try:
            if position is None:
                position = len(workbook.worksheets)
            
            ws = workbook.create_sheet(name, position)
            ws.sheet_properties.tabColor = tab_color
            
            logger.info(f"Custom worksheet created: {name}")
            return ws
            
        except Exception as e:
            logger.error(f"Error creating custom worksheet: {e}")
            raise
    
    def setup_worksheet_headers(self, worksheet: Worksheet, headers: List[str]):
        """
        Setup standard headers for a worksheet
        
        Args:
            worksheet: Excel worksheet
            headers: List of header names
        """
        try:
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                logger.debug(f"Header set: {header} at column {col}")
            
            logger.debug("Worksheet headers setup complete")
            
        except Exception as e:
            logger.error(f"Error setting up worksheet headers: {e}")
            raise
    
    def get_worksheet_config(self, name: str) -> Optional[Dict]:
        """
        Get configuration for a specific worksheet
        
        Args:
            name: Worksheet name
            
        Returns:
            Configuration dictionary or None
        """
        return self.worksheet_config.get(name)
    
    def validate_worksheet_structure(self, workbook: Workbook) -> List[str]:
        """
        Validate the worksheet structure against expected configuration
        
        Args:
            workbook: Excel workbook to validate
            
        Returns:
            List of validation issues (empty if valid)
        """
        issues = []
        
        try:
            expected_sheets = set(self.worksheet_config.keys())
            actual_sheets = set(workbook.sheetnames)
            
            # Check for missing worksheets
            missing = expected_sheets - actual_sheets
            if missing:
                issues.extend([f"Missing worksheet: {sheet}" for sheet in missing])
            
            # Check worksheet order
            for name, config in self.worksheet_config.items():
                if name in workbook.sheetnames:
                    actual_position = workbook.sheetnames.index(name)
                    if actual_position != config["position"]:
                        issues.append(f"Worksheet '{name}' at position {actual_position}, expected {config['position']}")
            
            if issues:
                logger.warning(f"Worksheet validation issues: {issues}")
            else:
                logger.info("Worksheet structure validation passed")
            
            return issues
            
        except Exception as e:
            logger.error(f"Error validating worksheet structure: {e}")
            return [f"Validation error: {e}"]
