"""
Template Loader for Personal Investment Analysis

This module provides Excel template loading and management capabilities
for standardized report generation.
"""

import logging
import os
from pathlib import Path
from typing import Dict, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


class TemplateLoader:
    """
    Template Loader for Excel Reports
    
    Manages loading and applying Excel templates for consistent
    report formatting and structure.
    """
    
    def __init__(self, template_dir: str = "templates"):
        """
        Initialize Template Loader
        
        Args:
            template_dir: Directory containing Excel templates
        """
        self.template_dir = Path(template_dir)
        self.template_dir.mkdir(exist_ok=True)
        
        self.available_templates = {
            "executive_summary": "executive_summary_template.xlsx",
            "detailed_analysis": "detailed_analysis_template.xlsx",
            "action_plan": "action_plan_template.xlsx",
            "risk_assessment": "risk_assessment_template.xlsx"
        }
        
        logger.info("Template Loader initialized")
    
    def load_template(self, template_name: str) -> Optional[Workbook]:
        """
        Load an Excel template
        
        Args:
            template_name: Name of the template to load
            
        Returns:
            Loaded workbook or None if template not found
        """
        try:
            if template_name not in self.available_templates:
                logger.warning(f"Template not found: {template_name}")
                return None
            
            template_file = self.template_dir / self.available_templates[template_name]
            
            if not template_file.exists():
                logger.warning(f"Template file not found: {template_file}")
                return self._create_default_template(template_name)
            
            workbook = load_workbook(template_file)
            logger.info(f"Template loaded: {template_name}")
            return workbook
            
        except Exception as e:
            logger.error(f"Error loading template {template_name}: {e}")
            return self._create_default_template(template_name)
    
    def create_template_structure(self, template_name: str) -> Workbook:
        """
        Create a new template structure
        
        Args:
            template_name: Name of the template to create
            
        Returns:
            New workbook with template structure
        """
        try:
            workbook = Workbook()
            
            if template_name == "executive_summary":
                self._setup_executive_summary_template(workbook)
            elif template_name == "detailed_analysis":
                self._setup_detailed_analysis_template(workbook)
            elif template_name == "action_plan":
                self._setup_action_plan_template(workbook)
            elif template_name == "risk_assessment":
                self._setup_risk_assessment_template(workbook)
            else:
                self._setup_generic_template(workbook)
            
            logger.info(f"Template structure created: {template_name}")
            return workbook
            
        except Exception as e:
            logger.error(f"Error creating template structure: {e}")
            raise
    
    def save_template(self, workbook: Workbook, template_name: str):
        """
        Save a workbook as a template
        
        Args:
            workbook: Workbook to save as template
            template_name: Name for the template
        """
        try:
            if template_name not in self.available_templates:
                logger.warning(f"Unknown template name: {template_name}")
                return
            
            template_file = self.template_dir / self.available_templates[template_name]
            workbook.save(template_file)
            
            logger.info(f"Template saved: {template_file}")
            
        except Exception as e:
            logger.error(f"Error saving template: {e}")
            raise
    
    def get_available_templates(self) -> Dict[str, str]:
        """
        Get list of available templates
        
        Returns:
            Dictionary of template names and their file paths
        """
        return self.available_templates.copy()
    
    def _create_default_template(self, template_name: str) -> Workbook:
        """Create a default template when file not found"""
        logger.info(f"Creating default template: {template_name}")
        return self.create_template_structure(template_name)
    
    def _setup_executive_summary_template(self, workbook: Workbook):
        """Setup executive summary template structure"""
        ws = workbook.active
        ws.title = "Executive Summary"
        
        # Basic structure
        ws['A1'] = "Investment Analysis - Executive Summary"
        ws['A3'] = "Report Generated:"
        ws['A4'] = "Total Recommendations:"
        ws['A5'] = "High Priority Actions:"
        ws['A7'] = "Key Insights"
        
        logger.debug("Executive summary template setup complete")
    
    def _setup_detailed_analysis_template(self, workbook: Workbook):
        """Setup detailed analysis template structure"""
        ws = workbook.active
        ws.title = "Detailed Analysis"
        
        # Headers
        headers = ["Category", "Priority", "Description", "Expected Impact", "Timeline"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        logger.debug("Detailed analysis template setup complete")
    
    def _setup_action_plan_template(self, workbook: Workbook):
        """Setup action plan template structure"""
        ws = workbook.active
        ws.title = "Action Plan"
        
        # Headers
        headers = ["Priority", "Action", "Category", "Timeline", "Status"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        logger.debug("Action plan template setup complete")
    
    def _setup_risk_assessment_template(self, workbook: Workbook):
        """Setup risk assessment template structure"""
        ws = workbook.active
        ws.title = "Risk Assessment"
        
        # Basic structure
        ws['A1'] = "Risk Assessment Summary"
        ws['A3'] = "Risk Recommendations Count:"
        ws['A5'] = "Risk Recommendations"
        
        # Headers for risk data
        headers = ["Priority", "Risk Type", "Description", "Mitigation"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=6, column=col, value=header)
        
        logger.debug("Risk assessment template setup complete")
    
    def _setup_generic_template(self, workbook: Workbook):
        """Setup generic template structure"""
        ws = workbook.active
        ws.title = "Report"
        
        ws['A1'] = "Investment Analysis Report"
        ws['A3'] = "Generated on:"
        
        logger.debug("Generic template setup complete")
