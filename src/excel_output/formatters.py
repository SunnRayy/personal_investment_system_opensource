"""
Excel Formatters for Personal Investment Analysis

This module provides Excel formatting and styling capabilities
for professional report generation.
"""

import logging
from typing import Optional
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, FORMAT_CURRENCY_USD

logger = logging.getLogger(__name__)


class ExcelFormatter:
    """
    Excel Formatter for Professional Styling
    
    Provides comprehensive formatting capabilities for Excel worksheets
    including fonts, colors, borders, and number formats.
    """
    
    def __init__(self):
        """Initialize Excel Formatter with standard styles"""
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.title_font = Font(bold=True, size=16)
        self.subtitle_font = Font(bold=True, size=14)
        self.body_font = Font(size=11)
        
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.accent_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
        
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.left_alignment = Alignment(horizontal='left', vertical='center')
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        logger.info("Excel Formatter initialized")
    
    def format_summary_sheet(self, worksheet: Worksheet):
        """
        Apply formatting to summary worksheets
        
        Args:
            worksheet: Excel worksheet to format
        """
        try:
            # Apply general formatting
            self._apply_general_formatting(worksheet)
            
            # Format title cells
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    if cell.value:
                        cell.font = self.title_font
                        cell.alignment = self.center_alignment
            
            # Auto-fit columns
            self._auto_fit_columns(worksheet)
            
            logger.debug("Summary sheet formatting applied")
            
        except Exception as e:
            logger.error(f"Error formatting summary sheet: {e}")
            raise
    
    def format_data_sheet(self, worksheet: Worksheet):
        """
        Apply formatting to data worksheets
        
        Args:
            worksheet: Excel worksheet to format
        """
        try:
            # Apply general formatting
            self._apply_general_formatting(worksheet)
            
            # Format header row
            for cell in worksheet[1]:
                if cell.value:
                    cell.font = self.header_font
                    cell.fill = self.header_fill
                    cell.alignment = self.center_alignment
                    cell.border = self.thin_border
            
            # Format data rows
            for row_num in range(2, worksheet.max_row + 1):
                for cell in worksheet[row_num]:
                    cell.font = self.body_font
                    cell.alignment = self.left_alignment
                    cell.border = self.thin_border
                    
                    # Alternate row coloring
                    if row_num % 2 == 0:
                        cell.fill = self.accent_fill
            
            # Auto-fit columns
            self._auto_fit_columns(worksheet)
            
            logger.debug("Data sheet formatting applied")
            
        except Exception as e:
            logger.error(f"Error formatting data sheet: {e}")
            raise
    
    def format_financial_cells(self, worksheet: Worksheet, cell_range: str):
        """
        Apply financial number formatting to specified range
        
        Args:
            worksheet: Excel worksheet
            cell_range: Range of cells to format (e.g., 'B2:B10')
        """
        try:
            for row in worksheet[cell_range]:
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = FORMAT_CURRENCY_USD
            
            logger.debug(f"Financial formatting applied to range: {cell_range}")
            
        except Exception as e:
            logger.error(f"Error applying financial formatting: {e}")
            raise
    
    def format_percentage_cells(self, worksheet: Worksheet, cell_range: str):
        """
        Apply percentage formatting to specified range
        
        Args:
            worksheet: Excel worksheet
            cell_range: Range of cells to format (e.g., 'C2:C10')
        """
        try:
            for row in worksheet[cell_range]:
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = FORMAT_PERCENTAGE
            
            logger.debug(f"Percentage formatting applied to range: {cell_range}")
            
        except Exception as e:
            logger.error(f"Error applying percentage formatting: {e}")
            raise
    
    def apply_priority_coloring(self, worksheet: Worksheet, priority_column: int, start_row: int = 2):
        """
        Apply color coding based on priority values
        
        Args:
            worksheet: Excel worksheet
            priority_column: Column number containing priority values
            start_row: Starting row for formatting
        """
        try:
            priority_colors = {
                "HIGH": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
                "MEDIUM": PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
                "LOW": PatternFill(start_color="95E1D3", end_color="95E1D3", fill_type="solid")
            }
            
            for row_num in range(start_row, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_num, column=priority_column)
                if cell.value and cell.value.upper() in priority_colors:
                    cell.fill = priority_colors[cell.value.upper()]
            
            logger.debug("Priority coloring applied")
            
        except Exception as e:
            logger.error(f"Error applying priority coloring: {e}")
            raise
    
    def create_title_section(self, worksheet: Worksheet, title: str, row: int = 1):
        """
        Create a formatted title section
        
        Args:
            worksheet: Excel worksheet
            title: Title text
            row: Row number for title
        """
        try:
            cell = worksheet.cell(row=row, column=1, value=title)
            cell.font = self.title_font
            cell.alignment = self.center_alignment
            
            # Merge cells across columns
            end_column = max(6, worksheet.max_column)
            worksheet.merge_cells(f"A{row}:{chr(64 + end_column)}{row}")
            
            logger.debug(f"Title section created: {title}")
            
        except Exception as e:
            logger.error(f"Error creating title section: {e}")
            raise
    
    def _apply_general_formatting(self, worksheet: Worksheet):
        """Apply general formatting to worksheet"""
        # Set default font for entire worksheet
        for row in worksheet.iter_rows():
            for cell in row:
                if not cell.font.name:
                    cell.font = self.body_font
    
    def _auto_fit_columns(self, worksheet: Worksheet):
        """Auto-fit column widths based on content"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = None
            
            # Find the first non-merged cell to get column letter
            for cell in column:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            # If we couldn't find a column letter, skip this column
            if not column_letter:
                continue
            
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter') and cell.value:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except Exception:
                    pass
            
            # Set column width with reasonable limits
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
