"""
Chart Builder for Personal Investment Analysis

This module provides chart and visualization capabilities
for Excel reports in investment analysis.
"""

import logging
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

logger = logging.getLogger(__name__)


class ChartBuilder:
    """
    Chart Builder for Excel Reports
    
    Creates professional charts and visualizations for investment
    analysis reports including pie charts, bar charts, and line charts.
    """
    
    def __init__(self):
        """Initialize Chart Builder"""
        self.chart_styles = {
            "pie": {
                "width": 15,
                "height": 10,
                "style": 10
            },
            "bar": {
                "width": 15,
                "height": 10,
                "style": 2
            },
            "line": {
                "width": 15,
                "height": 10,
                "style": 2
            }
        }
        
        logger.info("Chart Builder initialized")
    
    def create_priority_pie_chart(
        self,
        worksheet: Worksheet,
        data_range: str,
        position: str = "E2"
    ) -> PieChart:
        """
        Create a pie chart showing priority distribution
        
        Args:
            worksheet: Excel worksheet
            data_range: Range containing priority data
            position: Position to place chart
            
        Returns:
            Created pie chart
        """
        try:
            chart = PieChart()
            chart.title = "Priority Distribution"
            chart.style = self.chart_styles["pie"]["style"]
            chart.width = self.chart_styles["pie"]["width"]
            chart.height = self.chart_styles["pie"]["height"]
            
            # Add data
            data = Reference(worksheet, range_string=data_range)
            chart.add_data(data, titles_from_data=True)
            
            # Position the chart
            worksheet.add_chart(chart, position)
            
            logger.info("Priority pie chart created")
            return chart
            
        except Exception as e:
            logger.error(f"Error creating priority pie chart: {e}")
            raise
    
    def create_recommendations_bar_chart(
        self,
        worksheet: Worksheet,
        categories_range: str,
        values_range: str,
        position: str = "E2"
    ) -> BarChart:
        """
        Create a bar chart showing recommendations by category
        
        Args:
            worksheet: Excel worksheet
            categories_range: Range containing category labels
            values_range: Range containing values
            position: Position to place chart
            
        Returns:
            Created bar chart
        """
        try:
            chart = BarChart()
            chart.type = "col"
            chart.title = "Recommendations by Category"
            chart.style = self.chart_styles["bar"]["style"]
            chart.width = self.chart_styles["bar"]["width"]
            chart.height = self.chart_styles["bar"]["height"]
            
            # Add data
            data = Reference(worksheet, range_string=values_range)
            categories = Reference(worksheet, range_string=categories_range)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            # Styling
            chart.y_axis.title = "Number of Recommendations"
            chart.x_axis.title = "Category"
            
            # Position the chart
            worksheet.add_chart(chart, position)
            
            logger.info("Recommendations bar chart created")
            return chart
            
        except Exception as e:
            logger.error(f"Error creating recommendations bar chart: {e}")
            raise
    
    def create_risk_assessment_chart(
        self,
        worksheet: Worksheet,
        risk_data: Dict[str, float],
        position: str = "E2"
    ) -> PieChart:
        """
        Create a chart showing risk assessment breakdown
        
        Args:
            worksheet: Excel worksheet
            risk_data: Dictionary of risk categories and values
            position: Position to place chart
            
        Returns:
            Created risk chart
        """
        try:
            # First, add the risk data to the worksheet
            start_row = worksheet.max_row + 2
            
            for i, (category, value) in enumerate(risk_data.items()):
                worksheet.cell(row=start_row + i, column=1, value=category)
                worksheet.cell(row=start_row + i, column=2, value=value)
            
            # Create the chart
            chart = PieChart()
            chart.title = "Risk Assessment Breakdown"
            chart.style = self.chart_styles["pie"]["style"]
            chart.width = self.chart_styles["pie"]["width"]
            chart.height = self.chart_styles["pie"]["height"]
            
            # Define data ranges
            categories = Reference(worksheet, 
                                 min_col=1, min_row=start_row,
                                 max_col=1, max_row=start_row + len(risk_data) - 1)
            data = Reference(worksheet,
                           min_col=2, min_row=start_row,
                           max_col=2, max_row=start_row + len(risk_data) - 1)
            
            chart.add_data(data)
            chart.set_categories(categories)
            
            # Position the chart
            worksheet.add_chart(chart, position)
            
            logger.info("Risk assessment chart created")
            return chart
            
        except Exception as e:
            logger.error(f"Error creating risk assessment chart: {e}")
            raise
    
    def create_timeline_chart(
        self,
        worksheet: Worksheet,
        timeline_data: Dict[str, int],
        position: str = "E2"
    ) -> BarChart:
        """
        Create a chart showing action timeline distribution
        
        Args:
            worksheet: Excel worksheet
            timeline_data: Dictionary of timeframes and counts
            position: Position to place chart
            
        Returns:
            Created timeline chart
        """
        try:
            # Add timeline data to worksheet
            start_row = worksheet.max_row + 2
            
            for i, (timeframe, count) in enumerate(timeline_data.items()):
                worksheet.cell(row=start_row + i, column=1, value=timeframe)
                worksheet.cell(row=start_row + i, column=2, value=count)
            
            # Create the chart
            chart = BarChart()
            chart.type = "col"
            chart.title = "Action Timeline Distribution"
            chart.style = self.chart_styles["bar"]["style"]
            chart.width = self.chart_styles["bar"]["width"]
            chart.height = self.chart_styles["bar"]["height"]
            
            # Define data ranges
            categories = Reference(worksheet,
                                 min_col=1, min_row=start_row,
                                 max_col=1, max_row=start_row + len(timeline_data) - 1)
            data = Reference(worksheet,
                           min_col=2, min_row=start_row,
                           max_col=2, max_row=start_row + len(timeline_data) - 1)
            
            chart.add_data(data, titles_from_data=False)
            chart.set_categories(categories)
            
            # Styling
            chart.y_axis.title = "Number of Actions"
            chart.x_axis.title = "Timeline"
            
            # Position the chart
            worksheet.add_chart(chart, position)
            
            logger.info("Timeline chart created")
            return chart
            
        except Exception as e:
            logger.error(f"Error creating timeline chart: {e}")
            raise
    
    def add_chart_data_table(
        self,
        worksheet: Worksheet,
        data: Dict[str, any],
        start_position: Tuple[int, int] = (1, 1)
    ):
        """
        Add a data table to support chart creation
        
        Args:
            worksheet: Excel worksheet
            data: Dictionary of data to add
            start_position: Starting position (row, column)
        """
        try:
            start_row, start_col = start_position
            
            for i, (key, value) in enumerate(data.items()):
                worksheet.cell(row=start_row + i, column=start_col, value=key)
                worksheet.cell(row=start_row + i, column=start_col + 1, value=value)
            
            logger.debug("Chart data table added")
            
        except Exception as e:
            logger.error(f"Error adding chart data table: {e}")
            raise
    
    def apply_chart_styling(self, chart, title: str, colors: Optional[List[str]] = None):
        """
        Apply consistent styling to charts
        
        Args:
            chart: Chart object to style
            title: Chart title
            colors: Optional list of colors for chart elements
        """
        try:
            chart.title = title
            
            if colors and hasattr(chart, 'series') and chart.series:
                for i, series in enumerate(chart.series):
                    if i < len(colors):
                        # Apply color if supported
                        if hasattr(series, 'graphicalProperties'):
                            series.graphicalProperties.solidFill = colors[i]
            
            logger.debug(f"Chart styling applied: {title}")
            
        except Exception as e:
            logger.error(f"Error applying chart styling: {e}")
            raise
